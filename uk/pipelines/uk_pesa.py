"""
PESA — Public Expenditure Statistical Analyses 2025 (HM Treasury)
Source: HM Treasury PESA 2025
Licence: OGL v3
Updates: annual (July)
Requires: openpyxl  (pip install openpyxl)

Chapter 1: Total managed expenditure by function (what the money is spent ON)
Chapter 4: Departmental expenditure (who spends it)
"""

import duckdb
import httpx
import os
import tempfile
from datetime import datetime

DB_PATH = os.getenv("DB_PATH", "/opt/nstate/data/nstate.duckdb")

PESA_URLS = {
    "chapter1": "https://assets.publishing.service.gov.uk/media/6874fa8f92691289bdb7d394/PESA_2025_CP_Chapter_1_tables.xlsx",
    "chapter4": "https://assets.publishing.service.gov.uk/media/6874fe33b1b4ebc2c2e46574/PESA_2025_CP_Chapter_4_tables.xlsx",
}

# Years covered in PESA 2025 (outturn + plans)
PESA_YEARS = [
    2014,
    2015,
    2016,
    2017,
    2018,
    2019,
    2020,
    2021,
    2022,
    2023,
    2024,
    2025,
    2026,
    2027,
]


def download(url: str) -> str:
    r = httpx.get(url, timeout=120, follow_redirects=True)
    r.raise_for_status()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(r.content)
        return f.name


def parse_chapter1(path: str, db, source_id: str) -> int:
    """Parse functional expenditure — Table 1.1: TME by function."""
    try:
        import openpyxl
    except ImportError:
        print("  ERROR: openpyxl not installed.")
        return 0

    wb = openpyxl.load_workbook(path, data_only=True)
    loaded = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row with years
        year_row_idx = None
        years = []
        for i, row in enumerate(rows[:20]):
            vals = [str(v).strip() if v is not None else "" for v in row]
            year_vals = [
                v
                for v in vals
                if v.isdigit() and len(v) == 4 and 2010 <= int(v) <= 2030
            ]
            if len(year_vals) >= 5:
                years = [
                    int(v)
                    for v in vals
                    if v.isdigit() and len(v) == 4 and 2010 <= int(v) <= 2030
                ]
                year_row_idx = i
                year_cols = [
                    j
                    for j, v in enumerate(vals)
                    if v.isdigit() and len(v) == 4 and 2010 <= int(v) <= 2030
                ]
                break

        if not years or year_row_idx is None:
            continue

        for row in rows[year_row_idx + 1 :]:
            if not row or row[0] is None:
                continue
            function_name = str(row[0]).strip()
            if (
                not function_name
                or function_name.startswith("Note")
                or len(function_name) < 3
            ):
                continue

            for year, col_idx in zip(years, year_cols):
                val = row[col_idx] if col_idx < len(row) else None
                if val is None:
                    continue
                try:
                    amount = float(val)
                except (ValueError, TypeError):
                    continue
                try:
                    db.execute(
                        """
                        INSERT INTO uk_pesa_functional
                        (year, function_name, value_gbpm, source_id)
                        VALUES (?, ?, ?, ?)
                    """,
                        [year, function_name[:300], amount, source_id],
                    )
                    loaded += 1
                except Exception:
                    pass

    return loaded


def parse_chapter4(path: str, db, source_id: str) -> int:
    """Parse departmental expenditure — Table 4.x: DEL/AME by department."""
    try:
        import openpyxl
    except ImportError:
        return 0

    wb = openpyxl.load_workbook(path, data_only=True)
    loaded = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        year_row_idx = None
        years = []
        year_cols = []

        for i, row in enumerate(rows[:20]):
            vals = [str(v).strip() if v is not None else "" for v in row]
            yr = [
                v
                for v in vals
                if v.isdigit() and len(v) == 4 and 2010 <= int(v) <= 2030
            ]
            if len(yr) >= 5:
                years = [int(v) for v in yr]
                year_row_idx = i
                year_cols = [
                    j
                    for j, v in enumerate(vals)
                    if v.isdigit() and len(v) == 4 and 2010 <= int(v) <= 2030
                ]
                break

        if not years:
            continue

        for row in rows[year_row_idx + 1 :]:
            if not row or row[0] is None:
                continue
            dept_name = str(row[0]).strip()
            if not dept_name or dept_name.startswith("Note") or len(dept_name) < 3:
                continue

            for year, col_idx in zip(years, year_cols):
                val = row[col_idx] if col_idx < len(row) else None
                if val is None:
                    continue
                try:
                    amount = float(val)
                except (ValueError, TypeError):
                    continue
                try:
                    db.execute(
                        """
                        INSERT INTO uk_pesa_departmental
                        (year, department_name, expenditure_type, value_gbpm, source_id)
                        VALUES (?, ?, ?, ?, ?)
                    """,
                        [year, dept_name[:300], sheet_name[:100], amount, source_id],
                    )
                    loaded += 1
                except Exception:
                    pass

    return loaded


def run():
    print(f"[{datetime.utcnow().isoformat()}] Running PESA 2025 pipeline...")

    source_id = f"pesa_2025_{datetime.utcnow().strftime('%Y%m')}"

    with duckdb.connect(DB_PATH) as db:
        db.execute(
            """
            INSERT OR REPLACE INTO meta_sources
            (id, country, institution, dataset_name, release_url, licence, loaded_at)
            VALUES (?, 'uk', 'HM Treasury', 'PESA 2025',
                    'https://www.gov.uk/government/statistics/public-expenditure-statistical-analyses-2025',
                    'OGL v3', ?)
        """,
            [source_id, datetime.utcnow().isoformat()],
        )

        db.execute("DELETE FROM uk_pesa_functional WHERE source_id = ?", [source_id])
        db.execute("DELETE FROM uk_pesa_departmental WHERE source_id = ?", [source_id])

        total = 0
        for chapter, url in PESA_URLS.items():
            print(f"  Downloading PESA {chapter}...")
            try:
                path = download(url)
                if chapter == "chapter1":
                    n = parse_chapter1(path, db, source_id)
                else:
                    n = parse_chapter4(path, db, source_id)
                print(f"  ✓ {n:,} rows from {chapter}")
                total += n
                os.unlink(path)
            except Exception as e:
                print(f"  ✗ {chapter} failed: {e}")

        db.execute(
            """
            UPDATE meta_datasets SET last_loaded = ?, row_count = ?
            WHERE id = 'uk_pesa_functional'
        """,
            [datetime.utcnow().isoformat(), total],
        )

    print(f"  ✓ Total: {total:,} rows into uk_pesa_functional + uk_pesa_departmental")


if __name__ == "__main__":
    run()
